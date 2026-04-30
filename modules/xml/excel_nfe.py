from io import BytesIO

from openpyxl import load_workbook

from modules.xml.descricao import dividir_descricao, transformar_texto_descricao


def encontrar_colunas(ws):
    """Encontra os índices das colunas principais no cabeçalho."""
    header = {}
    for cell in ws[1]:
        if not cell.value:
            continue
        nome = str(cell.value).strip()
        chave = nome.split(":")[-1].lower()
        header[chave] = cell.column

    col_cprod = header.get("cprod")
    col_xprod = header.get("xprod")
    col_infad = header.get("infadprod")
    col_nitem = header.get("nitem")
    return col_nitem, col_cprod, col_xprod, col_infad


def processar_excel(excel_bytes):
    """Processa um Excel de espelho de NF-e, reordenando as descrições."""
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    col_nitem, col_cprod, col_xprod, col_infad = encontrar_colunas(ws)

    if col_cprod is None or col_xprod is None:
        return excel_bytes, []

    alteracoes = []

    for row in range(2, ws.max_row + 1):
        cprod = ws.cell(row=row, column=col_cprod).value
        xprod = ws.cell(row=row, column=col_xprod).value
        infad = ws.cell(row=row, column=col_infad).value if col_infad else ""
        nitem = ws.cell(row=row, column=col_nitem).value if col_nitem else row - 1

        if cprod is None or xprod is None:
            continue
        if str(cprod).strip() == "" or str(xprod).strip() == "":
            continue

        cprod = str(cprod)
        xprod = str(xprod)
        infad = "" if infad is None else str(infad)
        texto_antes = f"{xprod}{infad}"
        texto_depois = transformar_texto_descricao(texto_antes, cprod)

        if texto_depois != texto_antes:
            if col_infad:
                novo_xprod, novo_infad = dividir_descricao(texto_depois)
            else:
                novo_xprod, novo_infad = texto_depois, ""

            ws.cell(row=row, column=col_xprod).value = novo_xprod
            if col_infad:
                ws.cell(row=row, column=col_infad).value = novo_infad or None
            alteracoes.append({
                "nItem": str(nitem),
                "cProd": cprod,
                "xProd_antes": texto_antes[:120] + "..." if len(texto_antes) > 120 else texto_antes,
                "xProd_depois": texto_depois[:120] + "..." if len(texto_depois) > 120 else texto_depois,
            })

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), alteracoes
